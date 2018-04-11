import os 
import openpyxl

files = [f for f in os.listdir('.') if os.path.isfile(f)]
print(files)
i = 4
w=0
temp  = openpyxl.load_workbook('Template.xlsx')
tempws = temp.get_sheet_by_name('Sheet1')
searchtemplate = [ cell.value for i in tempws.iter_rows(min_row=1, max_row=1, min_col=1, max_col = tempws.max_column) for cell in i]
print(searchtemplate)
while True:
    count = 0
    i += 1
    #Open Worksheets
    try: wb = openpyxl.load_workbook(files[i])
    except: break
    tb = wb.get_sheet_names()
    print('Opening '+files[i])
    ws = wb.get_sheet_by_name('Transaction Register1')
    nopos=0
    notdisc=0

    #Defining the List of column headers in the Crappy Store Data (NO NEED TO TOUCH)
    search = [ cell.value for i in ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col = ws.max_column) for cell in i]
    
    #Search for Transaction No. (NO NEED TO TOUCH)
    searchtransno = search.index('Transaction No.') + 1
    
    #Search for Other Variables (EDITABLE)
    searchtranstype = search.index('Transaction Type')+1
    searchretsale = search.index('Sale Is Return Sale')+1
    searchstoreno = search.index('Store No.') + 1
    searchstaff = search.index('Staff ID') + 1
    searchdate = search.index('Date') + 1
    searchtime = search.index('Time') + 1
    searchvat = search.index('VAT Bus.Posting Group') + 1
    searchnet = search.index('Net Amount') + 1
    searchgross = search.index('Gross Amount') + 1
    searchpay = search.index('Payment') + 1
    searchdiscount = search.index('Discount Amount') + 1

    searchincome = search.index('Income/Exp. Amount') + 1
    searchcost = search.index('Cost Amount') + 1
    
    print('Gathering data from ' + files[i])
    #Creating a list of Transactions Numbers without the None (NO NEED TO ADD)
    trans = [ cell.value for i in ws.iter_rows(min_row=3, max_row= ws.max_row, min_col=searchtransno, max_col =searchtransno) for cell in i if cell.value is not None]
    
    #Creating the other columns based on number of transacions (EDITABLE)
    retsale =  [cell.value for i in ws.iter_rows(min_row=3, max_row= 2+len(trans), min_col= searchretsale ,max_col=searchretsale) for cell in i ]
    transtype =[cell.value for i in ws.iter_cols(min_row= 3, max_row= 2 + len(trans), min_col=searchtranstype, max_col=searchtranstype) for cell in i]  
    storeno =  [cell.value for i in ws.iter_rows(min_row=3, max_row= 2+len(trans), min_col= searchstoreno ,max_col=searchstoreno) for cell in i ]
    
    staff =  [cell.value for i in ws.iter_rows(min_row=3, max_row= 2+len(trans), min_col= searchstaff ,max_col=searchstaff) for cell in i ]
    date =[cell.value for i in ws.iter_cols(min_row= 3, max_row= 2 + len(trans), min_col=searchdate, max_col=searchdate) for cell in i]
    time = [cell.value for i in ws.iter_rows(min_row=3, max_row= 2+len(trans), min_col= searchtime ,max_col=searchtime) for cell in i ]
    vat =[cell.value for i in ws.iter_cols(min_row= 3, max_row= 2 + len(trans), min_col=searchvat, max_col=searchvat) for cell in i]
    net = [cell.value for i in ws.iter_rows(min_row=3, max_row= 2+len(trans), min_col= searchnet ,max_col=searchnet) for cell in i ]
    gross =[cell.value for i in ws.iter_cols(min_row= 3, max_row= 2 + len(trans), min_col=searchgross, max_col=searchgross) for cell in i]  
    pay = [cell.value for i in ws.iter_rows(min_row=3, max_row= 2+len(trans), min_col= searchpay ,max_col=searchpay) for cell in i ]
    discount =[cell.value for i in ws.iter_cols(min_row= 3, max_row= 2 + len(trans), min_col=searchdiscount, max_col=searchdiscount) for cell in i]
    
    income =[cell.value for i in ws.iter_cols(min_row= 3, max_row= 2 + len(trans), min_col=searchincome, max_col=searchincome) for cell in i]
    cost =[cell.value for i in ws.iter_cols(min_row= 3, max_row= 2 + len(trans), min_col=searchcost, max_col=searchcost) for cell in i]
    print('Data gathered from ' + files[i])
#==============================================================================
#     TEMPLATE AREA
#==============================================================================
     #Get Index for Defined Variables in the Template Sheet
    transnoindex = searchtemplate.index('Transaction No.') + 1
    
    transtypeindex = searchtemplate.index('Transaction Type') + 1
    retsaleindex = searchtemplate.index('Sale Is Return Sale')+1
    storenoindex = searchtemplate.index('Store No.') + 1

    staffindex = searchtemplate.index('Staff ID') + 1
    dateindex = searchtemplate.index('Date')+1
    timeindex = searchtemplate.index('Time') + 1
    vatindex = searchtemplate.index('VAT Bus.Posting Group')+1
    netindex = searchtemplate.index('Net Amount') + 1
    grossindex = searchtemplate.index('Gross Amount')+1
    payindex = searchtemplate.index('Payment') + 1
    discountindex = searchtemplate.index('Discount Amount')+1
    incomeindex = searchtemplate.index('Income/Exp. Amount')+1
    costindex = searchtemplate.index('Cost Amount')+1
    



    
    #Write in Template (EDITABLE)
    a1 = tempws.iter_cols(min_row=tempws.max_row+1, max_row = tempws.max_row + len(trans), min_col =transtypeindex, max_col = transtypeindex)
    a2 = tempws.iter_cols(min_row=tempws.max_row+1, max_row = tempws.max_row + len(trans), min_col =transnoindex, max_col = transnoindex)
    a3 = tempws.iter_cols(min_row=tempws.max_row+1, max_row = tempws.max_row + len(trans), min_col =retsaleindex, max_col = retsaleindex)
    a4 = tempws.iter_cols(min_row=tempws.max_row+1, max_row = tempws.max_row + len(trans), min_col=16, max_col = 16)
    a5 = tempws.iter_cols(min_row=tempws.max_row+1, max_row = tempws.max_row + len(trans), min_col =storenoindex, max_col = storenoindex)
    
     
    a7 = tempws.iter_cols(min_row=tempws.max_row+1, max_row = tempws.max_row + len(trans), min_col =staffindex, max_col = staffindex)
    a8 = tempws.iter_cols(min_row=tempws.max_row+1, max_row = tempws.max_row + len(trans), min_col =dateindex, max_col = dateindex)
    a9 = tempws.iter_cols(min_row=tempws.max_row+1, max_row = tempws.max_row + len(trans), min_col =timeindex, max_col = timeindex)
    a10 =tempws.iter_cols(min_row=tempws.max_row+1, max_row = tempws.max_row + len(trans), min_col =vatindex, max_col = vatindex)
    a11 =tempws.iter_cols(min_row=tempws.max_row+1, max_row = tempws.max_row + len(trans), min_col =netindex, max_col = netindex)
    a12 =tempws.iter_cols(min_row=tempws.max_row+1, max_row = tempws.max_row + len(trans), min_col =grossindex, max_col = grossindex)
    a13 =tempws.iter_cols(min_row=tempws.max_row+1, max_row = tempws.max_row + len(trans), min_col =payindex, max_col = payindex)
    a14 =tempws.iter_cols(min_row=tempws.max_row+1, max_row = tempws.max_row + len(trans), min_col =discountindex, max_col = discountindex)
    
    
    a16 =tempws.iter_cols(min_row=tempws.max_row+1, max_row = tempws.max_row + len(trans), min_col =incomeindex, max_col = incomeindex)
    a17 =tempws.iter_cols(min_row=tempws.max_row+1, max_row = tempws.max_row + len(trans), min_col =costindex, max_col = costindex)
    print('Writing in Output.xlsx')
    
    for transtypew,retsalew,transnow,filew,storenow,staffw,datew,timew,vatw,netw,grossw,payw,discountw,incomew,costw  in zip(a1,a2,a3,a4,a5,a7,a8,a9,a10,a11,a12,a13,a14,a16,a17):
        for a,b,c,d,e,g,h,r,j,k,l,m,n,p,q in zip(transtypew,retsalew,transnow,filew,storenow,staffw,datew,timew,vatw,netw,grossw,payw,discountw,incomew,costw):
            a.value = transtype[0]
            b.value = retsale[0]
            c.value = trans[0]
            d.value = files[i]
            e.value = storeno[0]
            
            g.value = staff[0]
            h.value = date[0]
            r.value = time[0]
            j.value = vat[0]
            k.value = net[0]
            l.value = gross[0]
            m.value = pay[0]
            n.value = discount[0]
            p.value = income[0]
            q.value = cost[0]
    #Deletion of the List one by one
            del transtype[0]
            del retsale[0]
            del trans[0]
            del storeno[0]
            
            del staff[0]
            del date[0]
            del time[0]
            del vat[0]
            del net[0]
            del gross[0]
            del pay[0]
            del discount[0]
            del income[0]
            del cost[0]
            count+=1
    print('A total of ' + str(count) + ' rows written\n')

 
    if i==52: break
    
temp.save('Output.xlsx')
print ('The file has been saved as Output.xlsx\n')
